import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import sys

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)  # when running as .exe
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # when running as .py

INPUT_PATH = os.path.join(BASE_DIR, "Employee_Availability.xlsx")

DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
SHIFTS = ['morning', 'evening', 'night']
HOURS_PER_SHIFT = 8
FULL_TIME_MIN = 70
FULL_TIME_MAX = 80
TOTAL_SHIFTS_PER_DAY = 3
EMPLOYEES_PER_SHIFT = 2

def read_schedule(filepath):
    df = pd.read_excel(filepath)
    schedule = {}
    for _, row in df.iterrows():
        name = row['Name']
        availability = []
        for day in DAYS:
            value = str(row[day]).strip().lower()
            if value in ['morning', 'evening', 'night']:
                availability.append([value])
            elif value in ['yes', 'all-day']:
                availability.append(['morning', 'evening', 'night'])
            elif value in ['morning/evening', 'evening/morning']:
                availability.append(['morning', 'evening'])
            elif value in ['morning/night', 'night/morning']:
                availability.append(['morning', 'night'])
            elif value in ['evening/night', 'night/evening']:
                availability.append(['evening', 'night'])
            elif value == 'no':
                availability.append([])
            else:
                availability.append([])  # fallback
        schedule[name] = availability
    return schedule

def is_valid(employee, day, shift, assignment, last_shift):
    if shift not in employee_availability[employee][day]:
        return False
    if day in assignment and any(employee in v for v in assignment[day].values()):
        return False  # already working this day
    last = last_shift.get(employee)
    #If worked night previous night wont assign morning to next day
    if last == 'night' and shift == 'morning':
        return False
    return True

def assign_shift(day, shift_index, assignment, hours, last_shift, used_oncall):
    if day == 7:
        return assignment

    shift = SHIFTS[shift_index]
    candidates = sorted(employee_availability.keys(), key=lambda x: hours[x])  # sort by least worked

    if shift_index == 0:
        assignment[day] = {}

    for i in range(len(candidates)):
        for j in range(i + 1, len(candidates)):
            e1, e2 = candidates[i], candidates[j]
            if all([
                is_valid(e1, day, shift, assignment, last_shift),
                is_valid(e2, day, shift, assignment, last_shift),
                hours[e1] + HOURS_PER_SHIFT <= FULL_TIME_MAX,
                hours[e2] + HOURS_PER_SHIFT <= FULL_TIME_MAX,
                not (used_oncall[e1] and used_oncall[e2] and shift_index == 0)
            ]):
                assignment[day][shift] = [e1, e2]
                hours[e1] += HOURS_PER_SHIFT
                hours[e2] += HOURS_PER_SHIFT
                last_shift[e1] = shift
                last_shift[e2] = shift

                next_shift_index = (shift_index + 1) % 3
                next_day = day + 1 if next_shift_index == 0 else day
                result = assign_shift(next_day, next_shift_index, assignment.copy(), hours.copy(), last_shift.copy(), used_oncall)
                if result:
                    return result
    return None

def save_schedule_to_excel(schedule, filename="Schedule.xlsx"):
    
    save_path = os.path.join(BASE_DIR, filename)


    # Create employee × day matrix
    output = {emp: {day: '' for day in DAYS} for emp in employee_availability.keys()}
    for day_idx, shifts in schedule.items():
        for shift, emps in shifts.items():
            for emp in emps:
                output[emp][DAYS[day_idx]] = shift.lower()

    df = pd.DataFrame.from_dict(output, orient='index')
    df.index.name = 'Employee'
    df.reset_index(inplace=True)

    df.to_excel(save_path, index=False)

    # === Apply formatting ===
    wb = load_workbook(save_path)
    ws = wb.active

    # Define fills
    fill_map = {
        'morning': PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid"),  # light yellow
        'evening': PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),  # light blue
        'night': PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid"),    # light red
        '': PatternFill(start_color="000000", end_color="000000", fill_type="solid"),         # black
    }

    white_font = Font(color="FFFFFF")

    # Apply colors
    for row in ws.iter_rows(min_row=2, min_col=2):  # skip header row and "Employee" column
        for cell in row:
            shift = str(cell.value).strip().lower()
            fill = fill_map.get(shift, fill_map[''])
            cell.fill = fill
            if shift == '':
                cell.font = white_font
            else:
                cell.value = shift.capitalize()

     # Auto-fit column widths
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2


    wb.save(save_path)

    

if __name__ == "__main__":
    employee_availability = read_schedule(INPUT_PATH)
    all_employees = list(employee_availability.keys())
    on_call = all_employees[-4:]
    used_oncall = {e: e in on_call for e in all_employees}

    assignment = {}
    hours = {e: 0 for e in all_employees}
    last_shift = {}

    result = assign_shift(0, 0, assignment, hours, last_shift, used_oncall)

    if result:
        save_schedule_to_excel(result)
        print("Schedule saved to 'schedule.xlsx'")
    else:
        print("Unable to find valid schedule.")

    input("\nPress Enter to exit...")


